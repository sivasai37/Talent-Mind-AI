from django.test import TestCase, RequestFactory

from ..views import RankingExportAPIView
from ..models import RankingJob, RankingResult, Candidate


class ExportTests(TestCase):
    def setUp(self):
        # create candidates and a ranking job
        self.c1 = Candidate.objects.create(full_name="A One")
        self.c2 = Candidate.objects.create(full_name="B Two")
        self.job = RankingJob.objects.create(title="Export Test", job_description="x")
        RankingResult.objects.create(ranking_job=self.job, candidate=self.c1, rank=1, semantic_score=80, skill_score=70, experience_score=60, recruitability_score=90, llm_score=50, final_score=75, strengths=["s"], weaknesses=["w"], missing_skills=["m"], recruiter_summary="ok", payload={})
        RankingResult.objects.create(ranking_job=self.job, candidate=self.c2, rank=2, semantic_score=60, skill_score=50, experience_score=40, recruitability_score=60, llm_score=30, final_score=45, strengths=["s2"], weaknesses=["w2"], missing_skills=["m2"], recruiter_summary="ok2", payload={})

    def test_export_latest_defaults(self):
        rf = RequestFactory()
        req = rf.get("/fake")
        resp = RankingExportAPIView.as_view()(req)
        # render and check
        if hasattr(resp, "render"):
            resp.render()
        content = resp.content.decode()
        self.assertIn("candidate_id,full_name,semantic_score", content)
        self.assertIn("A One", content)

    def test_excel_export(self):
        from ..views import RankingExcelExportAPIView
        import io
        import openpyxl

        # Test valid export
        rf = RequestFactory()
        req = rf.get(f"/api/export/{self.job.id}/")
        resp = RankingExcelExportAPIView.as_view()(req, ranking_id=self.job.id)

        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertEqual(
            resp["Content-Disposition"],
            'attachment; filename="submission.xlsx"',
        )

        # Read back Excel data
        excel_file = io.BytesIO(resp.content)
        wb = openpyxl.load_workbook(excel_file)
        self.assertIn("sample_submission", wb.sheetnames)

        ws = wb["sample_submission"]
        self.assertEqual(ws.max_row, 3)  # Header + 2 data rows

        # Check header
        headers = [ws.cell(row=1, column=col).value for col in range(1, 5)]
        self.assertEqual(headers, ["candidate_id", "rank", "score", "reasoning"])

        # Check bold font on headers
        for col in range(1, 5):
            self.assertTrue(ws.cell(row=1, column=col).font.bold)

        # Check values
        row2 = [ws.cell(row=2, column=col).value for col in range(1, 5)]
        self.assertEqual(row2[0], f"CAND_{self.c1.id:07d}")
        self.assertEqual(row2[1], 1)
        # score is normalized (max final_score is 75 -> 1.0, min is 45 -> 0.0)
        self.assertEqual(row2[2], 1.0)
        # reasoning fallback generated or checked
        self.assertTrue(row2[3])

        row3 = [ws.cell(row=3, column=col).value for col in range(1, 5)]
        self.assertEqual(row3[0], f"CAND_{self.c2.id:07d}")
        self.assertEqual(row3[1], 2)
        self.assertEqual(row3[2], 0.0)

        # Test non-existing job
        req_invalid = rf.get("/api/export/9999/")
        resp_invalid = RankingExcelExportAPIView.as_view()(req_invalid, ranking_id=9999)
        self.assertEqual(resp_invalid.status_code, 404)


