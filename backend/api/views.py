"""
views.py — HireGenius AI REST API Views

Endpoints:
  POST /api/analyze-jd/       — Analyze and structure a job description
  POST /api/rank/             — Rank candidates against a job description
  GET  /api/candidates/       — List all candidates
  GET  /api/rankings/         — List all ranking jobs
  GET  /api/rankings/export/  — Export rankings as CSV
  GET  /api/rankings/<id>/    — Get a specific ranking job
  POST /api/interview-questions/ — Generate interview questions for a candidate
  POST /api/talent-insights/  — Generate AI talent market insights
  GET  /api/stats/            — Dashboard statistics
"""
from __future__ import annotations

import csv
import io
import re
from io import StringIO

from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from .models import Candidate, RankingJob, RankingResult
from .serializers import CandidateSerializer, RankingJobSerializer, RankingJobListSerializer, RankingResultSerializer
from .services import ranking as ranking_service
from ranking_engine.jd import analyze_job_description
from django.conf import settings


# ---- Input sanitization helpers ----

def _sanitize_csv_field(value: str) -> str:
    """Prevent CSV injection by escaping dangerous leading characters."""
    if not value:
        return ""
    value = str(value)
    # Strip leading formula-injection characters
    if value and value[0] in ("=", "+", "-", "@", "\t", "\r"):
        value = "'" + value
    return value


def _validate_jd_input(data: dict) -> tuple[str, str, int]:
    """Validate and sanitize job description input. Returns (title, jd_text, top_k)."""
    max_jd_len = getattr(settings, "MAX_JD_LENGTH", 20000)
    max_top_k = getattr(settings, "MAX_TOP_K", 100)

    title = str(data.get("title", "") or "")[:255]
    job_description = str(data.get("job_description", "") or "")

    if not job_description.strip():
        raise ValueError("job_description is required and cannot be empty.")

    if len(job_description) > max_jd_len:
        job_description = job_description[:max_jd_len]

    try:
        top_k = max(1, min(int(data.get("top_k", 10)), max_top_k))
    except (ValueError, TypeError):
        top_k = 10

    return title, job_description, top_k


def _serialize_ranking_results(job: RankingJob, live_results: list) -> list:
    """Merge persisted DB rows with live ranking payload for API consumers."""
    db_rows = {r.rank: r for r in job.results.select_related("candidate").order_by("rank")}
    serialized = []
    for idx, item in enumerate(live_results, start=1):
        row = db_rows.get(idx)
        if row is not None:
            serialized.append(RankingResultSerializer(row).data)
            continue
        explanation = item.get("explanation") or {}
        serialized.append({
            "id": None,
            "candidate": {
                "id": item.get("candidate_id"),
                "full_name": item.get("full_name", ""),
            },
            "rank": idx,
            "semantic_score": item.get("semantic_score", 0.0),
            "skill_score": item.get("skill_score", 0.0),
            "experience_score": item.get("experience_score", 0.0),
            "recruitability_score": item.get("recruitability_score", 0.0),
            "llm_score": item.get("llm_score", 0.0),
            "final_score": item.get("final_score", 0.0),
            "potential_score": item.get("potential_score", 0.0),
            "confidence_score": item.get("confidence_score", 0.0),
            "risk_level": item.get("risk_level", "medium"),
            "strengths": item.get("strengths", []),
            "weaknesses": item.get("weaknesses", []),
            "missing_skills": item.get("missing_skills", []),
            "learning_path": item.get("learning_path", []),
            "interview_questions": item.get("interview_questions", []),
            "future_roles": item.get("future_roles", []),
            "growth_forecast": item.get("growth_forecast", {}),
            "salary_fit": item.get("salary_fit", {}),
            "why_selected": item.get("why_selected", ""),
            "why_rejected": item.get("why_rejected", ""),
            "role_fit_analysis": item.get("role_fit_analysis", ""),
            "recruiter_recommendation": item.get("recruiter_recommendation", ""),
            "recruiter_summary": item.get("recruiter_summary", ""),
            "payload": item,
        })
    return serialized


# ---- API Views ----

class RankAPIView(APIView):
    """POST /api/rank/  { title?, job_description, top_k? }"""

    def post(self, request):
        try:
            title, job_description, top_k = _validate_jd_input(request.data)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        jd_struct = analyze_job_description(job_description, title=title)
        jd_struct["title"] = title  # ensure title is in struct

        # Prefer precomputed FAISS when ready; never block API on full index build
        if not ranking_service.is_index_ready():
            ranking_service.ensure_index(auto_build=False)

        results = ranking_service.search_job(
            job_description,
            top_k=top_k,
            job_struct=jd_struct,
            title=title,
        )

        role_type = ranking_service.detect_role_type(jd_struct, title=title)
        weights_used = ranking_service.get_role_weights(role_type)

        with transaction.atomic():
            job = RankingJob.objects.create(
                title=title,
                job_description=job_description,
                job_analysis=jd_struct,
                role_type=role_type,
                weights_used=weights_used,
            )

            results_to_create = []
            for idx, item in enumerate(results, start=1):
                cid = item["candidate_id"]
                if isinstance(cid, str) and cid.startswith("CAND_"):
                    cid = int(cid.split("_")[1])
                candidate = Candidate.objects.filter(pk=cid).first()
                if candidate is None:
                    name = item.get("full_name") or f"Candidate {cid}"
                    candidate, _ = Candidate.objects.get_or_create(
                        full_name=name,
                        defaults={
                            "profile_text": item.get("why_selected", "")[:2000],
                            "skills": ", ".join(item.get("matching_skills", [])[:12]),
                            "years_experience": float(item.get("payload", {}).get("years_experience", 0) or 0),
                        },
                    )
                results_to_create.append(
                    RankingResult(
                        ranking_job=job,
                        candidate=candidate,
                        rank=idx,
                        semantic_score=item.get("semantic_score", 0.0),
                        skill_score=item.get("skill_score", 0.0),
                        experience_score=item.get("experience_score", 0.0),
                        recruitability_score=item.get("recruitability_score", 0.0),
                        llm_score=item.get("llm_score", 0.0),
                        final_score=item.get("final_score", 0.0),
                        potential_score=item.get("potential_score", 0.0),
                        confidence_score=item.get("confidence_score", 0.0),
                        risk_level=item.get("risk_level", "medium"),
                        strengths=item.get("strengths", []),
                        weaknesses=item.get("weaknesses", []),
                        missing_skills=item.get("missing_skills", []),
                        learning_path=item.get("learning_path", []),
                        interview_questions=item.get("interview_questions", []),
                        future_roles=item.get("future_roles", []),
                        growth_forecast=item.get("growth_forecast", {}),
                        salary_fit=item.get("salary_fit", {}),
                        why_selected=item.get("why_selected", ""),
                        why_rejected=item.get("why_rejected", ""),
                        role_fit_analysis=item.get("role_fit_analysis", ""),
                        recruiter_recommendation=item.get("recruiter_recommendation", "Moderate Fit"),
                        recruiter_summary=item.get("recruiter_summary", ""),
                        payload=item,
                    )
                )
            RankingResult.objects.bulk_create(results_to_create)

        serializer = RankingJobSerializer(job)
        response_data = serializer.data
        response_data["job_structure"] = jd_struct
        response_data["role_type"] = role_type
        response_data["weights_used"] = weights_used
        if len(response_data.get("results", [])) < len(results):
            response_data["results"] = _serialize_ranking_results(job, results)
        return Response(response_data, status=status.HTTP_200_OK)


class AnalyzeJDAPIView(APIView):
    """POST /api/analyze-jd/ { title?, job_description } -> structured JD"""

    def post(self, request):
        try:
            title, job_description, _ = _validate_jd_input(request.data)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        jd_struct = analyze_job_description(job_description, title=title)
        role_type = ranking_service.detect_role_type(jd_struct, title=title)
        jd_struct["role_type"] = role_type
        jd_struct["weights"] = ranking_service.get_role_weights(role_type)
        return Response(jd_struct, status=status.HTTP_200_OK)


class CandidateListAPIView(APIView):
    def get(self, request):
        candidates = Candidate.objects.all().order_by("full_name")[:100]
        serializer = CandidateSerializer(candidates, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class RankingListAPIView(APIView):
    def get(self, request):
        jobs = RankingJob.objects.order_by("-created_at")[:20]
        serializer = RankingJobListSerializer(jobs, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class RankingDetailAPIView(APIView):
    def get(self, request, pk):
        job = get_object_or_404(RankingJob, pk=pk)
        serializer = RankingJobSerializer(job)
        return Response(serializer.data, status=status.HTTP_200_OK)


class RankingExportAPIView(APIView):
    def get(self, request):
        job_id = request.query_params.get("job_id")
        export_format = request.query_params.get("format", "full")

        if not job_id:
            job = RankingJob.objects.order_by("-created_at").first()
            if job is None:
                return Response(
                    {"detail": "No ranking job available to export."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        else:
            job = get_object_or_404(RankingJob, pk=job_id)

        output = io.StringIO()
        writer = csv.writer(output)

        if export_format == "submission":
            writer.writerow([
                "rank", "candidate_id", "full_name", "final_score",
                "recommendation", "potential_score", "confidence_score",
                "risk_level", "strengths", "weaknesses", "missing_skills",
                "learning_path", "recruiter_summary",
            ])
            for result in job.results.order_by("rank"):
                writer.writerow([
                    result.rank,
                    f"CAND_{result.candidate.id:07d}",
                    _sanitize_csv_field(result.candidate.full_name),
                    round(result.final_score, 2),
                    _sanitize_csv_field(result.recruiter_recommendation),
                    round(result.potential_score, 2),
                    round(result.confidence_score, 2),
                    result.risk_level,
                    _sanitize_csv_field(" | ".join(result.strengths)),
                    _sanitize_csv_field(" | ".join(result.weaknesses)),
                    _sanitize_csv_field(" | ".join(result.missing_skills)),
                    _sanitize_csv_field(" | ".join(result.learning_path)),
                    _sanitize_csv_field(result.recruiter_summary),
                ])
        else:
            writer.writerow([
                "rank", "candidate_id", "full_name",
                "semantic_score", "skill_score", "experience_score",
                "recruitability_score", "llm_score", "final_score",
                "potential_score", "confidence_score", "risk_level",
                "recommendation", "strengths", "weaknesses",
                "missing_skills", "learning_path", "future_roles",
                "why_selected", "recruiter_summary",
            ])
            for result in job.results.order_by("rank"):
                writer.writerow([
                    result.rank,
                    f"CAND_{result.candidate.id:07d}",
                    _sanitize_csv_field(result.candidate.full_name),
                    round(result.semantic_score, 2),
                    round(result.skill_score, 2),
                    round(result.experience_score, 2),
                    round(result.recruitability_score, 2),
                    round(result.llm_score, 2),
                    round(result.final_score, 2),
                    round(result.potential_score, 2),
                    round(result.confidence_score, 2),
                    result.risk_level,
                    _sanitize_csv_field(result.recruiter_recommendation),
                    _sanitize_csv_field(" | ".join(result.strengths)),
                    _sanitize_csv_field(" | ".join(result.weaknesses)),
                    _sanitize_csv_field(" | ".join(result.missing_skills)),
                    _sanitize_csv_field(" | ".join(result.learning_path)),
                    _sanitize_csv_field(", ".join(result.future_roles)),
                    _sanitize_csv_field(result.why_selected),
                    _sanitize_csv_field(result.recruiter_summary),
                ])

        response = HttpResponse(output.getvalue(), content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="hiregenius_ranked_{job.id}.csv"'
        response["X-Content-Type-Options"] = "nosniff"
        return response


class DashboardStatsAPIView(APIView):
    """GET /api/stats/ — Executive dashboard statistics"""

    def get(self, request):
        total_candidates = Candidate.objects.count()
        total_rankings = RankingJob.objects.count()
        latest_job = RankingJob.objects.order_by("-created_at").first()

        stats = {
            "total_candidates": total_candidates,
            "total_rankings": total_rankings,
            "latest_job_title": latest_job.title if latest_job else None,
            "latest_job_date": latest_job.created_at.isoformat() if latest_job else None,
        }

        if latest_job:
            results = latest_job.results.order_by("rank")
            stats["top_candidate"] = None
            stats["score_distribution"] = {"strong": 0, "moderate": 0, "weak": 0}
            stats["risk_distribution"] = {"low": 0, "medium": 0, "high": 0}
            stats["avg_final_score"] = 0.0
            stats["avg_potential_score"] = 0.0

            if results.exists():
                top = results.first()
                stats["top_candidate"] = {
                    "name": top.candidate.full_name,
                    "score": round(top.final_score, 1),
                    "recommendation": top.recruiter_recommendation,
                    "potential_score": round(top.potential_score, 1),
                    "risk_level": top.risk_level,
                }
                all_results = list(results)
                for r in all_results:
                    rec = (r.recruiter_recommendation or "").lower()
                    if "strong" in rec:
                        stats["score_distribution"]["strong"] += 1
                    elif "moderate" in rec:
                        stats["score_distribution"]["moderate"] += 1
                    else:
                        stats["score_distribution"]["weak"] += 1

                    stats["risk_distribution"][r.risk_level] = (
                        stats["risk_distribution"].get(r.risk_level, 0) + 1
                    )

                scores = [r.final_score for r in all_results]
                potential_scores = [r.potential_score for r in all_results]
                stats["avg_final_score"] = round(sum(scores) / len(scores), 1)
                stats["avg_potential_score"] = round(sum(potential_scores) / len(potential_scores), 1)
                stats["ranked_count"] = len(all_results)

                # Top 5 for funnel
                stats["hiring_funnel"] = [
                    {
                        "rank": r.rank,
                        "name": r.candidate.full_name,
                        "score": round(r.final_score, 1),
                        "potential": round(r.potential_score, 1),
                        "recommendation": r.recruiter_recommendation,
                        "risk": r.risk_level,
                    }
                    for r in all_results[:5]
                ]

        return Response(stats, status=status.HTTP_200_OK)


class TalentInsightsAPIView(APIView):
    """POST /api/talent-insights/ { job_description, role_type? }"""

    def post(self, request):
        job_description = str(request.data.get("job_description", ""))[:5000]
        role_type = str(request.data.get("role_type", "general"))

        # Generate insights using available data
        all_candidates = Candidate.objects.all()
        total = all_candidates.count()
        avg_exp = 0.0
        if total > 0:
            avg_exp = sum(c.years_experience for c in all_candidates) / total

        insights = {
            "role_type": role_type,
            "market_insights": [
                f"Demand for {role_type} roles is consistently high in current market",
                "AI/ML skills are increasingly expected across all technical roles",
                "Remote-first culture drives larger candidate pool but higher competition",
            ],
            "skill_trends": [
                "Cloud-native development (AWS/GCP/Azure) is becoming baseline",
                "AI-augmented development workflows are differentiating factors",
                "System design and architecture skills command 20-30% salary premium",
            ],
            "hiring_recommendations": [
                "Prioritize candidates with demonstrated project impact over years of experience",
                "Assess learning velocity — fast learners outperform in rapidly evolving tech",
                "Consider potential score equally alongside current capability score",
                "Remote collaboration skills are now as critical as technical skills",
            ],
            "talent_pool_stats": {
                "total_candidates_in_pool": total,
                "average_experience_years": round(avg_exp, 1),
            },
        }

        return Response(insights, status=status.HTTP_200_OK)


class RankingExcelExportAPIView(APIView):
    """GET /api/export/<ranking_id>/ — Export specific ranking results as Excel (xlsx)"""

    def get(self, request, ranking_id):
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter

        job = get_object_or_404(RankingJob, pk=ranking_id)
        results = job.results.select_related("candidate").order_by("rank")

        wb = Workbook()
        ws = wb.active
        ws.title = "sample_submission"

        # Headers
        headers = ["candidate_id", "rank", "score", "reasoning"]
        ws.append(headers)

        # Bold headers
        bold_font = Font(bold=True)
        for col_num in range(1, len(headers) + 1):
            ws.cell(row=1, column=col_num).font = bold_font

        # Calculate min-max scaling bounds for score normalization
        scores = [r.final_score for r in results]
        min_score = min(scores) if scores else 0.0
        max_score = max(scores) if scores else 0.0
        span = max_score - min_score

        for r in results:
            cid = f"CAND_{r.candidate.id:07d}"
            rank_pos = r.rank

            if span > 1e-6:
                normalized_score = (r.final_score - min_score) / span
            else:
                normalized_score = 1.0

            # reasoning logic: use existing reasoning if available, otherwise generate
            reasoning = r.why_selected.strip()
            if not reasoning:
                cand = r.candidate
                first_job = cand.jobs.order_by('-id').first()
                role = first_job.title if first_job else 'Professional'
                years = cand.years_experience
                matched_skills = r.payload.get('matching_skills', []) if r.payload else []
                skills_str = ', '.join(matched_skills[:4])
                rate = cand.recruiter_response_rate
                reasoning = f"{role} with {years:.1f} yrs; matched {skills_str}; response rate {rate:.0f}%."

            # Ensure we round/write to 3 decimal places
            ws.append([cid, rank_pos, round(normalized_score, 3), reasoning])

        # Set number format for score column (3rd column)
        for row in range(2, len(results) + 2):
            ws.cell(row=row, column=3).number_format = '0.000'

        # Auto-size columns based on value lengths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="submission.xlsx"'
        response["X-Content-Type-Options"] = "nosniff"
        wb.save(response)
        return response

